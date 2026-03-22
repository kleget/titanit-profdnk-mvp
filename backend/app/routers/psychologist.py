from __future__ import annotations

import csv
import io
import json
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.config import settings
from app.db import get_db
from app.dependencies import get_optional_user, require_csrf_token, require_psychologist_or_admin
from app.models import (
    Answer,
    InviteLink,
    Submission,
    Test,
    TestSection,
    User,
    UserRole,
)
from app.schemas.formula_preview import FormulaPreviewPayloadSchema
from app.services.access_reminders import build_psychologist_access_reminder
from app.services.client_fields import normalize_client_fields_config
from app.services.formulas import FormulaError, evaluate_formula
from app.services.invite_links import (
    INVITE_LINK_STATE_ACTIVE,
    INVITE_LINK_STATE_EXPIRED,
    INVITE_LINK_STATE_UNKNOWN,
    invite_link_limit_text,
    invite_link_state,
    invite_link_state_label,
    is_invite_link_exhausted,
)
from app.services.killer_analytics import build_killer_analytics
from app.services.reports import build_docx_report, build_report_context, render_html_report
from app.services.test_changes import log_test_change
from app.services.tests import (
    custom_client_fields_from_flat_form,
    create_test_from_payload,
    export_test_config,
    formulas_from_flat_form,
    report_templates_from_flat_form,
    sections_from_flat_form,
)
from app.web import templates

router = APIRouter(tags=["psychologist"])


UPLOAD_DIR = Path(__file__).resolve().parents[1] / "static" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "_", value)
    return cleaned.strip("_")[:50] or "report"


def _ensure_test_access(test: Test, user: User) -> None:
    if user.role == UserRole.ADMIN:
        return
    if test.psychologist_id != user.id:
        raise HTTPException(status_code=404, detail="Test not found")


def _normalize_label(value: str) -> str:
    return " ".join(value.strip().split())


def _submission_invite_label(submission: Submission) -> str:
    extra = submission.client_extra_json or {}
    raw_value = extra.get("invite_label")
    if isinstance(raw_value, str):
        cleaned = _normalize_label(raw_value)
        if cleaned:
            return cleaned
    return "\u041e\u0441\u043d\u043e\u0432\u043d\u0430\u044f \u0441\u0441\u044b\u043b\u043a\u0430"


def _submission_invite_link_id(submission: Submission) -> int | None:
    extra = submission.client_extra_json or {}
    raw_value = extra.get("invite_link_id")
    if isinstance(raw_value, int):
        return raw_value
    if isinstance(raw_value, str) and raw_value.isdigit():
        return int(raw_value)
    return None


def _parse_usage_limit(raw_value: str) -> int | None:
    cleaned = raw_value.strip()
    if not cleaned:
        return None
    try:
        parsed = int(cleaned)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                "\u041b\u0438\u043c\u0438\u0442 "
                "\u043f\u0440\u043e\u0445\u043e\u0436\u0434\u0435\u043d\u0438\u0439 "
                "\u0434\u043e\u043b\u0436\u0435\u043d \u0431\u044b\u0442\u044c "
                "\u0446\u0435\u043b\u044b\u043c \u0447\u0438\u0441\u043b\u043e\u043c"
            ),
        ) from exc
    if parsed <= 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "\u041b\u0438\u043c\u0438\u0442 "
                "\u043f\u0440\u043e\u0445\u043e\u0436\u0434\u0435\u043d\u0438\u0439 "
                "\u0434\u043e\u043b\u0436\u0435\u043d \u0431\u044b\u0442\u044c "
                "\u0431\u043e\u043b\u044c\u0448\u0435 \u043d\u0443\u043b\u044f"
            ),
        )
    if parsed > 1_000_000:
        raise HTTPException(
            status_code=400,
            detail=(
                "\u041b\u0438\u043c\u0438\u0442 "
                "\u043f\u0440\u043e\u0445\u043e\u0436\u0434\u0435\u043d\u0438\u0439 "
                "\u0441\u043b\u0438\u0448\u043a\u043e\u043c \u0431\u043e\u043b\u044c\u0448\u043e\u0439"
            ),
        )
    return parsed


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _parse_link_datetime(raw_value: str, field_label: str) -> datetime | None:
    cleaned = raw_value.strip()
    if not cleaned:
        return None
    try:
        # datetime-local приходит без TZ: трактуем как UTC для предсказуемости на демо.
        parsed = datetime.fromisoformat(cleaned)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=(
                f"\u041d\u0435\u043a\u043e\u0440\u0440\u0435\u043a\u0442\u043d\u0430\u044f "
                f"\u0434\u0430\u0442\u0430: {field_label}"
            ),
        ) from exc
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _format_link_window(link: InviteLink) -> str:
    start_text = (
        link.start_at.strftime("%Y-%m-%d %H:%M")
        if link.start_at
        else "\u0441\u0440\u0430\u0437\u0443"
    )
    end_text = (
        link.expires_at.strftime("%Y-%m-%d %H:%M")
        if link.expires_at
        else "\u0431\u0435\u0437 \u0434\u0435\u0434\u043b\u0430\u0439\u043d\u0430"
    )
    return f"{start_text} -> {end_text}"


def _safe_float(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _format_float(value: float | None, *, precision: int = 2) -> str:
    if value is None:
        return "-"
    return f"{value:.{precision}f}"


def _build_invite_link_maps(test: Test) -> tuple[dict[int, InviteLink], dict[str, InviteLink]]:
    by_id: dict[int, InviteLink] = {}
    by_label: dict[str, InviteLink] = {}
    for link in test.invite_links:
        by_id[link.id] = link
        by_label[link.label] = link
    return by_id, by_label


def _submission_invite_state(submission: Submission, invite_links_by_id: dict[int, InviteLink]) -> str:
    invite_link_id = _submission_invite_link_id(submission)
    if invite_link_id is None:
        return INVITE_LINK_STATE_ACTIVE
    link = invite_links_by_id.get(invite_link_id)
    if not link:
        return INVITE_LINK_STATE_UNKNOWN
    return invite_link_state(link)


def _submission_invite_limit_text(
    submission: Submission, invite_links_by_id: dict[int, InviteLink]
) -> str:
    invite_link_id = _submission_invite_link_id(submission)
    if invite_link_id is None:
        return "\u0431\u0435\u0437 \u043b\u0438\u043c\u0438\u0442\u0430"
    link = invite_links_by_id.get(invite_link_id)
    if not link:
        return "-"
    return invite_link_limit_text(link)


def _submission_rows_for_template(
    submissions: list[Submission],
    invite_links_by_id: dict[int, InviteLink],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for submission in submissions:
        state = _submission_invite_state(submission, invite_links_by_id)
        rows.append(
            {
                "submission": submission,
                "invite_label": _submission_invite_label(submission),
                "invite_state": state,
                "invite_state_label": invite_link_state_label(state),
            }
        )
    return rows


def _build_invite_groups(
    submissions: list[Submission],
    invite_links_by_label: dict[str, InviteLink],
) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for sub in submissions:
        label = _submission_invite_label(sub)
        state = INVITE_LINK_STATE_ACTIVE
        limit_text = "\u0431\u0435\u0437 \u043b\u0438\u043c\u0438\u0442\u0430"
        if label != "\u041e\u0441\u043d\u043e\u0432\u043d\u0430\u044f \u0441\u0441\u044b\u043b\u043a\u0430":
            link = invite_links_by_label.get(label)
            if link:
                state = invite_link_state(link)
                limit_text = invite_link_limit_text(link)
            else:
                state = INVITE_LINK_STATE_UNKNOWN
                limit_text = "-"
        if label not in grouped:
            grouped[label] = {
                "label": label,
                "count": 0,
                "last_submitted_at": sub.submitted_at,
                "link_state": state,
                "link_state_label": invite_link_state_label(state),
                "link_limit_text": limit_text,
            }
        grouped[label]["count"] = int(grouped[label]["count"]) + 1
        if sub.submitted_at > grouped[label]["last_submitted_at"]:
            grouped[label]["last_submitted_at"] = sub.submitted_at
    return sorted(
        grouped.values(),
        key=lambda item: (int(item["count"]), item["last_submitted_at"]),
        reverse=True,
    )


def _clone_sections_payload(test: Test) -> list[dict]:
    payload: list[dict] = []
    for section in sorted(test.sections, key=lambda item: item.position):
        questions_payload: list[dict] = []
        for question in sorted(section.questions, key=lambda item: item.position):
            questions_payload.append(
                {
                    "key": question.key,
                    "text": question.text,
                    "question_type": question.question_type.value,
                    "required": question.required,
                    "options_json": question.options_json,
                    "min_value": question.min_value,
                    "max_value": question.max_value,
                    "weight": question.weight,
                    "visibility_condition": question.visibility_condition_json,
                }
            )
        if questions_payload:
            payload.append(
                {
                    "title": section.title,
                    "questions": questions_payload,
                    "visibility_condition": section.visibility_condition_json,
                }
            )
    return payload


def _clone_formulas_payload(test: Test) -> list[dict]:
    payload: list[dict] = []
    for formula in sorted(test.formulas, key=lambda item: item.position):
        payload.append(
            {
                "key": formula.key,
                "label": formula.label,
                "expression": formula.expression,
                "description": formula.description,
            }
        )
    return payload


def _build_campaign_comparison(
    submissions: list[Submission],
    invite_links_by_id: dict[int, InviteLink],
    invite_links_by_label: dict[str, InviteLink],
) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for submission in submissions:
        label = _submission_invite_label(submission)
        state = _submission_invite_state(submission, invite_links_by_id)
        metrics = submission.metrics_json or {}
        score_percent = _safe_float(metrics.get("score_percent"))
        if score_percent is None:
            total_score = _safe_float(metrics.get("total_score"))
            max_score = _safe_float(metrics.get("max_score"))
            if total_score is not None and max_score and max_score > 0:
                score_percent = round((total_score / max_score) * 100, 2)
        completion_percent = _safe_float(metrics.get("completion_percent"))

        if label not in grouped:
            limit_text = "\u0431\u0435\u0437 \u043b\u0438\u043c\u0438\u0442\u0430"
            if label != "\u041e\u0441\u043d\u043e\u0432\u043d\u0430\u044f \u0441\u0441\u044b\u043b\u043a\u0430":
                link = invite_links_by_label.get(label)
                if link is not None:
                    limit_text = invite_link_limit_text(link)
                else:
                    limit_text = "-"
            grouped[label] = {
                "label": label,
                "count": 0,
                "last_submitted_at": submission.submitted_at,
                "link_state": state,
                "link_state_label": invite_link_state_label(state),
                "link_limit_text": limit_text,
                "score_percent_sum": 0.0,
                "score_percent_count": 0,
                "completion_percent_sum": 0.0,
                "completion_percent_count": 0,
            }

        row = grouped[label]
        row["count"] = int(row["count"]) + 1
        if submission.submitted_at > row["last_submitted_at"]:
            row["last_submitted_at"] = submission.submitted_at

        if score_percent is not None:
            row["score_percent_sum"] = float(row["score_percent_sum"]) + score_percent
            row["score_percent_count"] = int(row["score_percent_count"]) + 1
        if completion_percent is not None:
            row["completion_percent_sum"] = float(row["completion_percent_sum"]) + completion_percent
            row["completion_percent_count"] = int(row["completion_percent_count"]) + 1

    result: list[dict[str, object]] = []
    for row in grouped.values():
        score_count = int(row["score_percent_count"])
        completion_count = int(row["completion_percent_count"])
        avg_score_percent = (
            round(float(row["score_percent_sum"]) / score_count, 2) if score_count > 0 else None
        )
        avg_completion_percent = (
            round(float(row["completion_percent_sum"]) / completion_count, 2)
            if completion_count > 0
            else None
        )
        result.append(
            {
                "label": row["label"],
                "count": row["count"],
                "last_submitted_at": row["last_submitted_at"],
                "link_state": row["link_state"],
                "link_state_label": row["link_state_label"],
                "link_limit_text": row["link_limit_text"],
                "avg_score_percent": avg_score_percent,
                "avg_score_percent_text": _format_float(avg_score_percent),
                "avg_completion_percent": avg_completion_percent,
                "avg_completion_percent_text": _format_float(avg_completion_percent),
            }
        )

    return sorted(
        result,
        key=lambda item: (int(item["count"]), item["last_submitted_at"]),
        reverse=True,
    )


def _submission_score_percent(submission: Submission) -> float | None:
    metrics = submission.metrics_json or {}
    score_percent = _safe_float(metrics.get("score_percent"))
    if score_percent is not None:
        return round(score_percent, 2)
    total_score = _safe_float(metrics.get("total_score"))
    max_score = _safe_float(metrics.get("max_score"))
    if total_score is None or max_score is None or max_score <= 0:
        return None
    return round((total_score / max_score) * 100, 2)


def _build_submission_csv(
    test: Test,
    submissions: list[Submission],
    invite_links_by_id: dict[int, InviteLink],
) -> bytes:
    formula_keys: list[str] = []
    for formula in sorted(test.formulas, key=lambda item: item.position):
        key = formula.key.strip()
        if key and key not in formula_keys:
            formula_keys.append(key)

    header = [
        "submission_id",
        "submitted_at",
        "client_full_name",
        "client_email",
        "client_phone",
        "source_label",
        "source_state",
        "total_score",
        "max_score",
        "score_percent",
        "completion_percent",
    ]
    header.extend(formula_keys)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)

    for submission in submissions:
        metrics = submission.metrics_json or {}
        row = [
            submission.id,
            submission.submitted_at.isoformat(),
            submission.client_full_name,
            submission.client_email or "",
            submission.client_phone or "",
            _submission_invite_label(submission),
            _submission_invite_state(submission, invite_links_by_id),
            _format_float(_safe_float(metrics.get("total_score"))),
            _format_float(_safe_float(metrics.get("max_score"))),
            _format_float(_submission_score_percent(submission)),
            _format_float(_safe_float(metrics.get("completion_percent"))),
        ]
        for key in formula_keys:
            row.append(_format_float(_safe_float(metrics.get(key))))
        writer.writerow(row)

    return buffer.getvalue().encode("utf-8-sig")


def _build_submission_xlsx(
    test: Test,
    submissions: list[Submission],
    invite_links_by_id: dict[int, InviteLink],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    formula_keys: list[str] = []
    for formula in sorted(test.formulas, key=lambda item: item.position):
        key = formula.key.strip()
        if key and key not in formula_keys:
            formula_keys.append(key)

    header = [
        "submission_id",
        "submitted_at",
        "client_full_name",
        "client_email",
        "client_phone",
        "source_label",
        "source_state",
        "total_score",
        "max_score",
        "score_percent",
        "completion_percent",
    ]
    header.extend(formula_keys)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Submissions"
    sheet.append(header)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for index, _value in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill

    for submission in submissions:
        metrics = submission.metrics_json or {}
        row = [
            submission.id,
            submission.submitted_at.isoformat(),
            submission.client_full_name,
            submission.client_email or "",
            submission.client_phone or "",
            _submission_invite_label(submission),
            _submission_invite_state(submission, invite_links_by_id),
            _safe_float(metrics.get("total_score")),
            _safe_float(metrics.get("max_score")),
            _submission_score_percent(submission),
            _safe_float(metrics.get("completion_percent")),
        ]
        for key in formula_keys:
            row.append(_safe_float(metrics.get(key)))
        sheet.append(row)

    column_widths: dict[int, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell_value = "" if cell.value is None else str(cell.value)
            column_widths[cell.column] = max(column_widths.get(cell.column, 0), len(cell_value))
    for column_index, width in column_widths.items():
        adjusted = min(max(width + 2, 12), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = adjusted

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def _resolve_pdf_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        ("DejaVuSans", Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")),
        ("Arial", Path("/usr/share/fonts/truetype/msttcorefonts/Arial.ttf")),
        ("Arial", Path("C:/Windows/Fonts/arial.ttf")),
    ]
    for font_name, font_path in candidates:
        try:
            if not font_path.exists():
                continue
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
            return font_name
        except Exception:
            continue
    return "Helvetica"


def _build_campaign_pdf_lines(
    *,
    test: Test,
    campaign_comparison: list[dict[str, object]],
    analytics: dict[str, object],
) -> list[str]:
    lines = [
        "Платформа ПрофДНК — сводный аналитический отчет по кампаниям",
        f"Тест: {test.title}",
        f"Сформировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "Раздел 1. Сравнение кампаний",
    ]
    if not campaign_comparison:
        lines.append("Нет данных: пока нет прохождений по кампаниям.")
    else:
        lines.append("Метка | Прохождений | Ср.% балла | Ср.% заполнения | Статус")
        for item in campaign_comparison:
            line = (
                f"{item.get('label', '-')} | {item.get('count', 0)} | "
                f"{item.get('avg_score_percent_text', '-')} | "
                f"{item.get('avg_completion_percent_text', '-')} | "
                f"{item.get('link_state_label', '-')}"
            )
            lines.append(line[:160])

    lines.extend(
        [
            "",
            "Раздел 2. Динамика по источникам и датам",
        ]
    )
    source_dynamics = analytics.get("source_dynamics", {}) if isinstance(analytics, dict) else {}
    timeline_rows = (
        source_dynamics.get("timeline_rows", []) if isinstance(source_dynamics, dict) else []
    )
    if not timeline_rows:
        lines.append("Нет данных для динамики.")
    else:
        lines.append("Дата | Источник | Прохождений | Ср.% балла | Ср.% заполнения")
        for row in timeline_rows[:40]:
            line = (
                f"{row.get('date', '-')} | {row.get('source_label', '-')} | {row.get('count', 0)} | "
                f"{_format_float(_safe_float(row.get('avg_score_percent')))} | "
                f"{_format_float(_safe_float(row.get('avg_completion_percent')))}"
            )
            lines.append(line[:160])
        if len(timeline_rows) > 40:
            lines.append(f"... еще строк: {len(timeline_rows) - 40}")

    lines.extend(
        [
            "",
            "Раздел 3. Риски и аномалии",
        ]
    )
    risk_overview = analytics.get("risk_overview", {}) if isinstance(analytics, dict) else {}
    zone_counters = risk_overview.get("zone_counters", {}) if isinstance(risk_overview, dict) else {}
    lines.append(
        f"Высокий риск: {zone_counters.get('high', 0)} | "
        f"Средний риск: {zone_counters.get('medium', 0)} | "
        f"Низкий риск: {zone_counters.get('low', 0)}"
    )
    anomalies = analytics.get("anomalies", []) if isinstance(analytics, dict) else []
    if anomalies:
        lines.append("Клиенты с аномалиями:")
        for row in anomalies[:20]:
            flags = ", ".join(row.get("flags", [])[:2])
            lines.append(f"- {row.get('client_full_name', '-')}: {flags}"[:160])
        if len(anomalies) > 20:
            lines.append(f"... еще аномалий: {len(anomalies) - 20}")
    else:
        lines.append("Аномалии не обнаружены.")
    return lines


def _escape_pdf_text(text: str) -> str:
    ascii_text = text.encode("ascii", "replace").decode("ascii")
    return ascii_text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_simple_pdf(lines: list[str]) -> bytes:
    rendered = lines[:90]
    content_lines = ["BT", "/F1 11 Tf", "36 806 Td"]
    first_line = True
    for line in rendered:
        if not first_line:
            content_lines.append("0 -14 Td")
        first_line = False
        content_lines.append(f"({_escape_pdf_text(line)}) Tj")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", "replace")

    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>",
        b"<< /Length "
        + str(len(stream)).encode("ascii")
        + b" >>\nstream\n"
        + stream
        + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]

    result = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(result))
        result.extend(f"{index} 0 obj\n".encode("ascii"))
        result.extend(obj)
        result.extend(b"\nendobj\n")

    xref_offset = len(result)
    result.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    result.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        result.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    result.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF"
        ).encode("ascii")
    )
    return bytes(result)


def _build_campaign_pdf(
    *,
    test: Test,
    campaign_comparison: list[dict[str, object]],
    analytics: dict[str, object],
) -> bytes:
    lines = _build_campaign_pdf_lines(
        test=test,
        campaign_comparison=campaign_comparison,
        analytics=analytics,
    )
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError:
        return _build_simple_pdf(lines)

    pdf_buffer = io.BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=A4)
    _width, height = A4
    font_name = _resolve_pdf_font()
    y = height - 48

    def add_line(text: str) -> None:
        nonlocal y
        if not text:
            y -= 8
            return
        size = 10
        step = 14
        if text.startswith("Платформа ПрофДНК"):
            size, step = 14, 22
        elif text.startswith("Раздел "):
            size, step = 12, 18
        elif text.startswith("- "):
            size, step = 9, 12
        if y <= 48:
            pdf.showPage()
            y = height - 48
        pdf.setFont(font_name, size)
        pdf.drawString(36, y, text)
        y -= step

    for line in lines:
        add_line(line)

    pdf.save()
    pdf_buffer.seek(0)
    return pdf_buffer.read()

def _generate_unique_invite_token(db: Session) -> str:
    while True:
        token = secrets.token_urlsafe(24)
        invite_exists = db.scalar(select(InviteLink.id).where(InviteLink.token == token))
        test_exists = db.scalar(select(Test.id).where(Test.share_token == token))
        if not invite_exists and not test_exists:
            return token


@router.get("/")
def root_redirect(user: User | None = Depends(get_optional_user)) -> object:
    if not user:
        return RedirectResponse("/login", status_code=303)
    return RedirectResponse("/admin" if user.role == UserRole.ADMIN else "/dashboard", status_code=303)


@router.get("/dashboard")
def dashboard(
    request: Request,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        return RedirectResponse("/admin", status_code=303)

    tests_count = db.scalar(select(func.count(Test.id)).where(Test.psychologist_id == current_user.id)) or 0
    submissions_count = (
        db.scalar(
            select(func.count(Submission.id))
            .join(Test, Submission.test_id == Test.id)
            .where(Test.psychologist_id == current_user.id)
        )
        or 0
    )
    access_reminder = build_psychologist_access_reminder(current_user)
    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "title": "\u041b\u0438\u0447\u043d\u044b\u0439 \u043a\u0430\u0431\u0438\u043d\u0435\u0442",
            "user": current_user,
            "tests_count": tests_count,
            "submissions_count": submissions_count,
            "access_reminder": access_reminder,
            "base_url": settings.base_url,
        },
    )


@router.post("/profile")
def update_profile(
    about_md: str = Form(""),
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Admin profile is not editable here")
    current_user.about_md = about_md
    db.commit()
    return RedirectResponse("/dashboard?notice=profile_saved&notice_type=success", status_code=303)


@router.post("/profile/photo")
async def upload_photo(
    photo: UploadFile,
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Admin photo upload is disabled")
    if not photo.filename:
        raise HTTPException(status_code=400, detail="File is empty")
    suffix = Path(photo.filename).suffix.lower()
    if suffix not in {".png", ".jpg", ".jpeg", ".webp"}:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    file_name = f"{current_user.id}_{secrets.token_hex(6)}{suffix}"
    file_path = UPLOAD_DIR / file_name
    file_path.write_bytes(await photo.read())
    current_user.photo_filename = file_name
    db.commit()
    return RedirectResponse("/dashboard?notice=photo_uploaded&notice_type=success", status_code=303)


@router.get("/tests")
def tests_page(
    request: Request,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    query = select(Test).order_by(Test.created_at.desc()).options(selectinload(Test.submissions))
    if current_user.role == UserRole.PSYCHOLOGIST:
        query = query.where(Test.psychologist_id == current_user.id)
    tests = db.scalars(query).all()
    return templates.TemplateResponse(
        request,
        "tests.html",
        {
            "title": "\u041c\u043e\u0438 \u043e\u043f\u0440\u043e\u0441\u043d\u0438\u043a\u0438",
            "user": current_user,
            "tests": tests,
            "base_url": settings.base_url,
        },
    )


@router.post("/tests/{test_id}/clone")
def clone_test(
    test_id: int,
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin cannot clone tests")

    source_test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.sections).selectinload(TestSection.questions),
            selectinload(Test.formulas),
        )
    )
    if not source_test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(source_test, current_user)

    fields_config = normalize_client_fields_config(source_test.required_client_fields)
    cloned_test = create_test_from_payload(
        db=db,
        psychologist_id=current_user.id,
        title=f"{source_test.title} (копия)",
        description=source_test.description,
        allow_client_report=source_test.allow_client_report,
        required_client_fields=list(fields_config["required_builtin_fields"]),
        custom_client_fields=list(fields_config["custom_fields"]),
        report_templates=dict(fields_config["report_templates"]),
        sections_payload=_clone_sections_payload(source_test),
        formulas_payload=_clone_formulas_payload(source_test),
    )
    log_test_change(
        db,
        test_id=cloned_test.id,
        action="test_cloned",
        actor_user_id=current_user.id,
        details={"source_test_id": source_test.id},
    )
    db.commit()
    return RedirectResponse(
        f"/tests/{cloned_test.id}?notice=test_cloned&notice_type=success",
        status_code=303,
    )


@router.get("/tests/new")
def new_test_page(
    request: Request,
    current_user: User = Depends(require_psychologist_or_admin),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin cannot create tests")
    return templates.TemplateResponse(
        request,
        "test_builder.html",
        {
            "title": (
                "\u041a\u043e\u043d\u0441\u0442\u0440\u0443\u043a\u0442\u043e\u0440 "
                "\u043c\u0435\u0442\u043e\u0434\u0438\u043a"
            ),
            "user": current_user,
        },
    )


@router.post("/tests/new/manual")
async def create_test_manual(
    request: Request,
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin cannot create tests")

    form = await request.form()
    title = str(form.get("title", ""))
    description = str(form.get("description", ""))
    allow_client_report = str(form.get("allow_client_report", "false")).lower() == "true"
    required_client_fields = form.getlist("required_client_fields")
    custom_client_fields = custom_client_fields_from_flat_form(
        field_keys=form.getlist("cf_key[]"),
        field_labels=form.getlist("cf_label[]"),
        field_types=form.getlist("cf_type[]"),
        field_required=form.getlist("cf_required[]"),
        field_placeholders=form.getlist("cf_placeholder[]"),
    )
    report_templates = report_templates_from_flat_form(
        client_blocks=form.getlist("rt_client[]"),
        psychologist_blocks=form.getlist("rt_psychologist[]"),
    )

    sections = sections_from_flat_form(
        section_titles=form.getlist("section_titles[]"),
        question_texts=form.getlist("q_text[]"),
        question_keys=form.getlist("q_key[]"),
        question_types=form.getlist("q_type[]"),
        question_required=form.getlist("q_required[]"),
        question_options=form.getlist("q_options[]"),
        question_correct=form.getlist("q_correct[]"),
        question_min=form.getlist("q_min[]"),
        question_max=form.getlist("q_max[]"),
        question_weight=form.getlist("q_weight[]"),
        question_section_titles=form.getlist("q_section[]"),
        section_if_key=form.getlist("section_if_key[]"),
        section_if_operator=form.getlist("section_if_operator[]"),
        section_if_value=form.getlist("section_if_value[]"),
        question_if_key=form.getlist("q_if_key[]"),
        question_if_operator=form.getlist("q_if_operator[]"),
        question_if_value=form.getlist("q_if_value[]"),
    )
    formulas = formulas_from_flat_form(
        metric_keys=form.getlist("metric_key[]"),
        metric_labels=form.getlist("metric_label[]"),
        metric_expressions=form.getlist("metric_expression[]"),
        metric_descriptions=form.getlist("metric_description[]"),
    )
    test = create_test_from_payload(
        db=db,
        psychologist_id=current_user.id,
        title=title,
        description=description,
        allow_client_report=allow_client_report,
        required_client_fields=required_client_fields,
        custom_client_fields=custom_client_fields,
        report_templates=report_templates,
        sections_payload=sections,
        formulas_payload=formulas,
    )
    log_test_change(
        db,
        test_id=test.id,
        action="test_created_manual",
        actor_user_id=current_user.id,
        details={"sections": len(sections), "formulas": len(formulas)},
    )
    db.commit()
    return RedirectResponse(
        f"/tests/{test.id}?notice=test_created&notice_type=success",
        status_code=303,
    )


@router.post("/tests/new/import")
def create_test_import(
    title: str = Form(""),
    description: str = Form(""),
    allow_client_report: str = Form("true"),
    required_client_fields: list[str] = Form(default=[]),
    config_json: str = Form(...),
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    if current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin cannot create tests")
    try:
        parsed = json.loads(config_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON config") from exc
    sections = parsed.get("sections")
    if not isinstance(sections, list):
        raise HTTPException(status_code=400, detail="JSON must include sections[]")
    formulas = parsed.get("formula_metrics") or parsed.get("metric_formulas") or []
    if not isinstance(formulas, list):
        raise HTTPException(status_code=400, detail="JSON formula_metrics must be an array")
    parsed_client_fields = parsed.get("client_fields")
    if not isinstance(parsed_client_fields, dict):
        parsed_client_fields = {}
    custom_client_fields = (
        parsed.get("custom_client_fields")
        or parsed_client_fields.get("custom_fields")
        or []
    )
    if not isinstance(custom_client_fields, list):
        raise HTTPException(status_code=400, detail="JSON custom_client_fields must be an array")
    report_templates = parsed.get("report_templates")
    if report_templates is not None and not isinstance(report_templates, dict):
        raise HTTPException(status_code=400, detail="JSON report_templates must be an object")

    test = create_test_from_payload(
        db=db,
        psychologist_id=current_user.id,
        title=(title or parsed.get("title") or "").strip(),
        description=(description or parsed.get("description") or "").strip(),
        allow_client_report=allow_client_report.lower() == "true",
        required_client_fields=required_client_fields
        or parsed.get("required_client_fields")
        or parsed_client_fields.get("required_builtin_fields")
        or ["full_name"],
        custom_client_fields=custom_client_fields,
        report_templates=report_templates,
        sections_payload=sections,
        formulas_payload=formulas,
    )
    log_test_change(
        db,
        test_id=test.id,
        action="test_created_import",
        actor_user_id=current_user.id,
        details={"sections": len(sections), "formulas": len(formulas)},
    )
    db.commit()
    return RedirectResponse(
        f"/tests/{test.id}?notice=test_imported&notice_type=success",
        status_code=303,
    )


@router.get("/tests/{test_id}")
def test_detail(
    test_id: int,
    request: Request,
    source_status: str = "all",
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.psychologist),
            selectinload(Test.sections).selectinload(TestSection.questions),
            selectinload(Test.formulas),
            selectinload(Test.invite_links),
            selectinload(Test.submissions).selectinload(Submission.answers),
            selectinload(Test.change_logs),
        )
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    links_were_updated = False
    for link in test.invite_links:
        state = invite_link_state(link)
        if link.is_active and (
            is_invite_link_exhausted(link) or state == INVITE_LINK_STATE_EXPIRED
        ):
            link.is_active = False
            links_were_updated = True
    if links_were_updated:
        db.commit()

    submissions = sorted(test.submissions, key=lambda item: item.submitted_at, reverse=True)
    invite_links_by_id, invite_links_by_label = _build_invite_link_maps(test)
    invite_groups = _build_invite_groups(submissions, invite_links_by_label)
    campaign_comparison = _build_campaign_comparison(
        submissions,
        invite_links_by_id,
        invite_links_by_label,
    )
    killer_analytics = build_killer_analytics(test, submissions, invite_links_by_id)
    submission_rows = _submission_rows_for_template(submissions, invite_links_by_id)
    allowed_filters = {"all", "active", "pending", "expired", "exhausted", "disabled", "unknown"}
    if source_status not in allowed_filters:
        source_status = "all"
    if source_status != "all":
        submission_rows = [row for row in submission_rows if row["invite_state"] == source_status]

    invite_link_rows = []
    for link in test.invite_links:
        state = invite_link_state(link)
        invite_link_rows.append(
            {
                "link": link,
                "state": state,
                "state_label": invite_link_state_label(state),
                "limit_text": invite_link_limit_text(link),
                "window_text": _format_link_window(link),
                "single_use_text": "\u0414\u0430" if link.single_use else "\u041d\u0435\u0442",
                "target_client_name": link.target_client_name or "-",
            }
        )
    latest_change_logs = sorted(test.change_logs, key=lambda row: row.created_at, reverse=True)[:10]

    return templates.TemplateResponse(
        request,
        "test_detail.html",
        {
            "title": f"\u0422\u0435\u0441\u0442: {test.title}",
            "user": current_user,
            "test": test,
            "submission_rows": submission_rows,
            "invite_groups": invite_groups,
            "campaign_comparison": campaign_comparison,
            "killer_analytics": killer_analytics,
            "invite_link_rows": invite_link_rows,
            "source_status_filter": source_status,
            "latest_change_logs": latest_change_logs,
            "base_url": settings.base_url,
        },
    )


@router.post("/tests/{test_id}/links")
def create_invite_link(
    test_id: int,
    label: str = Form(""),
    usage_limit: str = Form(""),
    start_at: str = Form(""),
    expires_at: str = Form(""),
    single_use: str = Form("false"),
    target_client_name: str = Form(""),
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    test = db.scalar(select(Test).where(Test.id == test_id))
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    cleaned_label = _normalize_label(label)
    if not cleaned_label:
        raise HTTPException(status_code=400, detail="Link label is required")
    if len(cleaned_label) > 120:
        raise HTTPException(status_code=400, detail="Link label is too long")

    duplicate = db.scalar(
        select(InviteLink.id).where(InviteLink.test_id == test_id, InviteLink.label == cleaned_label)
    )
    if duplicate:
        raise HTTPException(status_code=400, detail="Link label already exists")

    parsed_usage_limit = _parse_usage_limit(usage_limit)
    parsed_start_at = _parse_link_datetime(start_at, "start_at")
    parsed_expires_at = _parse_link_datetime(expires_at, "expires_at")
    if (
        parsed_start_at is not None
        and parsed_expires_at is not None
        and parsed_expires_at <= parsed_start_at
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "\u0414\u0435\u0434\u043b\u0430\u0439\u043d "
                "\u0441\u0441\u044b\u043b\u043a\u0438 \u0434\u043e\u043b\u0436\u0435\u043d "
                "\u0431\u044b\u0442\u044c \u043f\u043e\u0437\u0436\u0435 "
                "\u0432\u0440\u0435\u043c\u0435\u043d\u0438 \u0441\u0442\u0430\u0440\u0442\u0430"
            ),
        )
    is_single_use = single_use.strip().lower() in {"true", "1", "yes", "on"}
    cleaned_target_client_name = _normalize_label(target_client_name)
    if cleaned_target_client_name and len(cleaned_target_client_name) > 255:
        raise HTTPException(
            status_code=400,
            detail=(
                "\u0418\u043c\u044f \u043a\u043b\u0438\u0435\u043d\u0442\u0430 "
                "\u0434\u043b\u044f \u043f\u0435\u0440\u0441\u043e\u043d\u0430\u043b\u044c\u043d\u043e\u0439 "
                "\u0441\u0441\u044b\u043b\u043a\u0438 \u0441\u043b\u0438\u0448\u043a\u043e\u043c "
                "\u0434\u043b\u0438\u043d\u043d\u043e\u0435"
            ),
        )
    if cleaned_target_client_name and not is_single_use:
        raise HTTPException(
            status_code=400,
            detail=(
                "\u041f\u0435\u0440\u0441\u043e\u043d\u0430\u043b\u044c\u043d\u0430\u044f "
                "\u0441\u0441\u044b\u043b\u043a\u0430 \u0434\u043e\u043b\u0436\u043d\u0430 "
                "\u0431\u044b\u0442\u044c \u043e\u0434\u043d\u043e\u0440\u0430\u0437\u043e\u0432\u043e\u0439"
            ),
        )

    invite_link = InviteLink(
        test_id=test.id,
        label=cleaned_label,
        token=_generate_unique_invite_token(db),
        is_active=True,
        usage_limit=parsed_usage_limit,
        start_at=parsed_start_at,
        expires_at=parsed_expires_at,
        single_use=is_single_use,
        target_client_name=cleaned_target_client_name or None,
    )
    db.add(invite_link)
    log_test_change(
        db,
        test_id=test.id,
        action="invite_link_created",
        actor_user_id=current_user.id,
        details={
            "label": cleaned_label,
            "usage_limit": parsed_usage_limit,
            "start_at": parsed_start_at.isoformat() if parsed_start_at else None,
            "expires_at": parsed_expires_at.isoformat() if parsed_expires_at else None,
            "single_use": is_single_use,
            "target_client_name": cleaned_target_client_name or None,
        },
    )
    db.commit()
    return RedirectResponse(
        f"/tests/{test.id}?notice=invite_link_added&notice_type=success",
        status_code=303,
    )


@router.post("/tests/{test_id}/links/{link_id}/toggle")
def toggle_invite_link(
    test_id: int,
    link_id: int,
    _: None = Depends(require_csrf_token),
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> object:
    test = db.scalar(select(Test).where(Test.id == test_id))
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    link = db.scalar(
        select(InviteLink).where(InviteLink.id == link_id, InviteLink.test_id == test_id)
    )
    if not link:
        raise HTTPException(status_code=404, detail="Invite link not found")

    if not link.is_active:
        state = invite_link_state(link)
        if is_invite_link_exhausted(link):
            raise HTTPException(
                status_code=400,
                detail=(
                    "\u0421\u0441\u044b\u043b\u043a\u0430 "
                    "\u0438\u0441\u0447\u0435\u0440\u043f\u0430\u043b\u0430 "
                    "\u043b\u0438\u043c\u0438\u0442 \u043f\u0440\u043e\u0445\u043e\u0436\u0434\u0435\u043d\u0438\u0439"
                ),
            )
        if state == INVITE_LINK_STATE_EXPIRED:
            raise HTTPException(
                status_code=400,
                detail=(
                    "\u0421\u0440\u043e\u043a "
                    "\u0434\u0435\u0439\u0441\u0442\u0432\u0438\u044f \u0441\u0441\u044b\u043b\u043a\u0438 "
                    "\u0438\u0441\u0442\u0451\u043a"
                ),
            )

    link.is_active = not link.is_active
    log_test_change(
        db,
        test_id=test.id,
        action="invite_link_toggled",
        actor_user_id=current_user.id,
        details={"label": link.label, "is_active": link.is_active},
    )
    db.commit()
    return RedirectResponse(
        f"/tests/{test.id}?notice=invite_link_toggled&notice_type=success",
        status_code=303,
    )


@router.post("/api/formulas/preview")
def formula_preview(
    payload: FormulaPreviewPayloadSchema,
    _: None = Depends(require_csrf_token),
    _current_user: User = Depends(require_psychologist_or_admin),
) -> JSONResponse:
    parsed = payload.to_service_payload()
    formulas_payload = parsed["formulas"]
    context: dict[str, float] = dict(parsed["context"])

    results: list[dict[str, object]] = []
    for formula in formulas_payload:
        key = str(formula["key"])
        label = str(formula["label"])
        expression = str(formula["expression"])
        try:
            value = evaluate_formula(expression, context)
            rounded = round(float(value), 4)
            context[key] = rounded
            results.append(
                {
                    "key": key,
                    "label": label,
                    "expression": expression,
                    "status": "ok",
                    "value": rounded,
                }
            )
        except FormulaError as exc:
            results.append(
                {
                    "key": key,
                    "label": label,
                    "expression": expression,
                    "status": "error",
                    "error": str(exc),
                }
            )

    return JSONResponse(
        {
            "results": results,
            "context": {name: round(number, 4) for name, number in context.items()},
        }
    )


@router.get("/tests/{test_id}/submissions.json")
def submissions_json(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> JSONResponse:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(selectinload(Test.submissions), selectinload(Test.invite_links))
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    links_were_updated = False
    for link in test.invite_links:
        state = invite_link_state(link)
        if link.is_active and (
            is_invite_link_exhausted(link) or state == INVITE_LINK_STATE_EXPIRED
        ):
            link.is_active = False
            links_were_updated = True
    if links_were_updated:
        db.commit()

    invite_links_by_id, _invite_links_by_label = _build_invite_link_maps(test)
    payload = []
    for sub in sorted(test.submissions, key=lambda item: item.submitted_at, reverse=True):
        state = _submission_invite_state(sub, invite_links_by_id)
        payload.append(
            {
                "id": sub.id,
                "client_full_name": sub.client_full_name,
                "submitted_at": sub.submitted_at.isoformat(),
                "total_score": _safe_float((sub.metrics_json or {}).get("total_score")),
                "max_score": _safe_float((sub.metrics_json or {}).get("max_score")),
                "score_percent": _submission_score_percent(sub),
                "completion_percent": _safe_float((sub.metrics_json or {}).get("completion_percent")),
                "invite_label": _submission_invite_label(sub),
                "invite_state": state,
                "invite_state_label": invite_link_state_label(state),
                "invite_limit_text": _submission_invite_limit_text(sub, invite_links_by_id),
            }
        )
    return JSONResponse(payload)


@router.get("/tests/{test_id}/submissions.csv")
def submissions_csv(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.submissions),
            selectinload(Test.invite_links),
            selectinload(Test.formulas),
        )
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    submissions = sorted(test.submissions, key=lambda item: item.submitted_at, reverse=True)
    invite_links_by_id, _invite_links_by_label = _build_invite_link_maps(test)
    csv_bytes = _build_submission_csv(test, submissions, invite_links_by_id)
    filename = _slugify(f"{test.title}_submissions_{test.id}") + ".csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers=headers,
    )


@router.get("/tests/{test_id}/submissions.xlsx")
def submissions_xlsx(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.submissions),
            selectinload(Test.invite_links),
            selectinload(Test.formulas),
        )
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    submissions = sorted(test.submissions, key=lambda item: item.submitted_at, reverse=True)
    invite_links_by_id, _invite_links_by_label = _build_invite_link_maps(test)
    xlsx_bytes = _build_submission_xlsx(test, submissions, invite_links_by_id)
    filename = _slugify(f"{test.title}_submissions_{test.id}") + ".xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/tests/{test_id}/campaign-report.pdf")
def campaign_report_pdf(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.submissions),
            selectinload(Test.invite_links),
            selectinload(Test.formulas),
        )
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    submissions = sorted(test.submissions, key=lambda item: item.submitted_at, reverse=True)
    invite_links_by_id, invite_links_by_label = _build_invite_link_maps(test)
    campaign_comparison = _build_campaign_comparison(
        submissions,
        invite_links_by_id,
        invite_links_by_label,
    )
    killer_analytics = build_killer_analytics(test, submissions, invite_links_by_id)
    pdf_bytes = _build_campaign_pdf(
        test=test,
        campaign_comparison=campaign_comparison,
        analytics=killer_analytics,
    )
    filename = _slugify(f"{test.title}_campaign_report_{test.id}") + ".pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers=headers,
    )


@router.get("/tests/{test_id}/export.json")
def export_test(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> JSONResponse:
    test = db.scalar(
        select(Test)
        .where(Test.id == test_id)
        .options(
            selectinload(Test.sections).selectinload(TestSection.questions),
            selectinload(Test.formulas),
        )
    )
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)
    return JSONResponse(export_test_config(test))


@router.get("/reports/{submission_id}/{report_kind}.html", response_class=HTMLResponse)
def report_html(
    submission_id: int,
    report_kind: str,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> HTMLResponse:
    if report_kind not in {"client", "psychologist"}:
        raise HTTPException(status_code=400, detail="Invalid report type")
    submission = db.scalar(
        select(Submission)
        .where(Submission.id == submission_id)
        .options(
            selectinload(Submission.answers).selectinload(Answer.question),
            selectinload(Submission.test)
            .selectinload(Test.sections)
            .selectinload(TestSection.questions),
            selectinload(Submission.test).selectinload(Test.psychologist),
        )
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _ensure_test_access(submission.test, current_user)
    if report_kind == "client" and not submission.test.allow_client_report:
        raise HTTPException(status_code=403, detail="Client report disabled")

    context = build_report_context(submission.test, submission, report_kind=report_kind)  # type: ignore[arg-type]
    html = render_html_report(context, report_kind=report_kind)  # type: ignore[arg-type]
    return HTMLResponse(content=html)


@router.get("/reports/{submission_id}/{report_kind}.docx")
def report_docx(
    submission_id: int,
    report_kind: str,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    if report_kind not in {"client", "psychologist"}:
        raise HTTPException(status_code=400, detail="Invalid report type")
    submission = db.scalar(
        select(Submission)
        .where(Submission.id == submission_id)
        .options(
            selectinload(Submission.answers).selectinload(Answer.question),
            selectinload(Submission.test)
            .selectinload(Test.sections)
            .selectinload(TestSection.questions),
            selectinload(Submission.test).selectinload(Test.psychologist),
        )
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    _ensure_test_access(submission.test, current_user)
    if report_kind == "client" and not submission.test.allow_client_report:
        raise HTTPException(status_code=403, detail="Client report disabled")

    context = build_report_context(submission.test, submission, report_kind=report_kind)  # type: ignore[arg-type]
    document = build_docx_report(context, report_kind=report_kind)  # type: ignore[arg-type]
    filename = _slugify(f"{submission.client_full_name}_{submission.id}_{report_kind}") + ".docx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        document,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers=headers,
    )


@router.get("/tests/{test_id}/stats")
def test_stats(
    test_id: int,
    current_user: User = Depends(require_psychologist_or_admin),
    db: Session = Depends(get_db),
) -> JSONResponse:
    test = db.scalar(select(Test).where(Test.id == test_id))
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    _ensure_test_access(test, current_user)

    submissions_count = db.scalar(select(func.count(Submission.id)).where(Submission.test_id == test_id)) or 0
    latest_submission = db.scalar(
        select(Submission)
        .where(Submission.test_id == test_id)
        .order_by(Submission.submitted_at.desc())
        .limit(1)
    )
    return JSONResponse(
        {
            "submissions_count": submissions_count,
            "last_submitted_at": latest_submission.submitted_at.isoformat() if latest_submission else None,
        }
    )

